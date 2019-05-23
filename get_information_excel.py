from function.aws_ec2 import AWSEC2
from function.aws_rds import AWSRDS
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# BPM
# instanceids = ['i-0f8b3df82161baf5f', 'i-03a1c9ed8eb9b637e', 'i-00e50946ad570a835']
instanceids = []

infomation_dict = {}
instances_dict = {}
securitygroup_dict = {}
ebs_dict = {}
route_tables_dict = {}

filters = [
    {
        'Name': 'tag:System',
        'Values': [
            'Branded Goods',
        ]
    },
    # {
    #     'Name': 'tag:System type',
    #     'Values': [
    #         'UAT',
    #     ]
    # },
]
filename = ''
for filter_dict in filters:
    filename = filename + filter_dict['Values'][0] + '_'


class GetInfo(object):
    def __init__(self):
        self.line_count = 0
        self.ec2_client = AWSEC2()
        self.rds_client = AWSRDS()
        self.instanceids = instanceids + self.ec2_client.get_instanceids(filters)
        # self.instanceids = instanceids
        self.sgids = set()
        # self.sgids = ['sg-099bc01e23c472901', ]
        self.subnetids = set()
        self.ebsids = []

    def get_ec2_info(self):
        for i in range(len(self.instanceids)):
            instanceid = self.instanceids.pop()
            instance_info = self.ec2_client.ec2_instance_describe(instanceid)['Instances'][0]
            instance_id = instance_info['InstanceId']
            instance_name = ''
            for tag in instance_info['Tags']:
                if tag['Key'] == 'Name':
                    instance_name = tag['Value']
            instance_type = instance_info['InstanceType']
            cpu_count = int(instance_info['CpuOptions']['CoreCount']) * int(
                instance_info['CpuOptions']['ThreadsPerCore'])
            ami_id = instance_info['ImageId']
            az = instance_info['Placement']['AvailabilityZone']
            vpcid = instance_info['VpcId']
            subnet_id = instance_info['SubnetId']
            self.subnetids.add(subnet_id)
            private_ip = instance_info['PrivateIpAddress']
            if 'PublicIpAddress' in instance_info.keys():
                public_ip = instance_info['PublicIpAddress']
            else:
                public_ip = None
            sg_ids = ''
            sg_infos = instance_info['NetworkInterfaces'][0]['Groups']
            for sg_info in sg_infos:
                sg_id = sg_info['GroupId']
                self.sgids.add(sg_id)
                sg_ids = sg_ids + sg_id + ';'
            key_name = instance_info['KeyName']
            ebs_ids = ''
            ebs_infos = instance_info['BlockDeviceMappings']
            for ebs_info in ebs_infos:
                ebs_id = ebs_info['Ebs']['VolumeId']
                self.ebsids.append(ebs_id)
                ebs_ids = ebs_ids + ebs_id + ';'
            instancs_lists = [instance_id, instance_name, instance_type, ebs_ids, cpu_count, ami_id, az, vpcid,
                              subnet_id,
                              private_ip, public_ip, sg_ids, key_name]
            for j in range(len(instancs_lists)):
                row = chr(97 + j) + str(2 + i)
                instances_dict[row] = instancs_lists[j]

    def get_volumes_info(self):
        for i in range(len(self.ebsids)):
            ebsid = self.ebsids.pop()
            volume_info = self.ec2_client.ec2_volume_describe(ebsid)
            instance_id = volume_info['Attachments'][0]['InstanceId']
            device_name = volume_info['Attachments'][0]['Device']
            volume_size = volume_info['Size']
            az = volume_info['AvailabilityZone']
            volume_type = volume_info['VolumeType']
            volume_lists = [ebsid, volume_type, instance_id, device_name, volume_size, az]
            for j in range(len(volume_lists)):
                row = chr(97 + j) + str(2 + i)
                ebs_dict[row] = volume_lists[j]

    def get_bound_info(self, bound_infos, num):
        line_num = 0
        count = 0
        for i in range(len(bound_infos)):
            bound_info = bound_infos[i]
            protocol = bound_info['IpProtocol']
            if protocol == '-1':
                protocol = 'All'
            port_range = ''
            if 'FromPort' not in bound_info.keys():
                port_range = 'All'
            elif bound_info['FromPort'] == -1:
                port_range = 'N/A'
            else:
                from_port = str(bound_info['FromPort'])
                to_port = str(bound_info['ToPort'])
                if from_port == to_port:
                    port_range = port_range + from_port
                else:
                    port_range = port_range + from_port + '-' + to_port
            ip_range = bound_info['IpRanges']
            grouppairs = bound_info['UserIdGroupPairs']
            line_num = line_num + len(ip_range) + len(grouppairs)
            if len(ip_range) != 0:
                for k in range(len(ip_range)):
                    source_info = ip_range[k]
                    source = source_info['CidrIp']
                    description = ''
                    if 'Description' in source_info.keys():
                        description = source_info['Description']
                    sg_lists = [protocol, port_range, source, description]
                    for j in range(len(sg_lists)):
                        row = chr(num + j) + str(count + 3 + self.line_count)
                        securitygroup_dict[row] = sg_lists[j]
                    count += 1
            if len(grouppairs) != 0:
                for l in range(len(grouppairs)):
                    group_pair = grouppairs[l]
                    groupid = group_pair['GroupId']
                    description = ''
                    if 'Description' in group_pair.keys():
                        description = group_pair['Description']
                    sg_lists = [protocol, port_range, groupid, description]
                    for j in range(len(sg_lists)):
                        row = chr(num + j) + str(count + 3 + self.line_count)
                        securitygroup_dict[row] = sg_lists[j]
                    count += 1
        return line_num

    def get_sg_info(self):
        sgs_info = self.ec2_client.ec2_security_groups_describe(filters)
        for i in range(len(sgs_info)):
            sg_info = sgs_info[i]
            sgid = sg_info['GroupId']
            # print(sg_info)
            sg_name = sg_info['GroupName']
            # inbound
            inbound_infos = sg_info['IpPermissions']
            num1 = self.get_bound_info(inbound_infos, 99)
            # outbound
            outbound_infos = sg_info['IpPermissionsEgress']
            num2 = self.get_bound_info(outbound_infos, 103)
            sg_lists = [sgid, sg_name]
            # print(num1, num2)
            if num1 > num2:
                dict_num = num1
            else:
                dict_num = num2
            for k in range(dict_num):
                for j in range(len(sg_lists)):
                    row = chr(97 + j) + str(3 + self.line_count + k)
                    securitygroup_dict[row] = sg_lists[j]
            self.line_count += dict_num

        return

    def get_vpc_info(self):
        rows = ['a', 'c']
        vpcs_info = self.ec2_client.ec2_vpcs_describe(filters)
        for i in range(len(vpcs_info)):
            vpcid = vpcs_info[i]['VpcId']
            cidr_block = vpcs_info[i]['CidrBlock']
            row_vpc = rows[0] + str(4 + i)
            rowcidr = rows[1] + str(4 + i)
            infomation_dict[row_vpc] = vpcid
            infomation_dict[rowcidr] = cidr_block

    def get_subnet_info(self):
        subnets_info = self.ec2_client.ec2_subnets_describe(filters)
        for i in range(len(subnets_info)):
            subnet_info = subnets_info[i]
            subnetid = subnet_info['SubnetId']
            tags = subnet_info['Tags']
            subnet_name = ''
            for tag in tags:
                if 'Name' == tag['Key']:
                    subnet_name = tag['Value']
                    break
            cidr = subnet_info['CidrBlock']
            az = subnet_info['AvailabilityZone']
            azid = subnet_info['AvailabilityZoneId']
            subnet_list = [subnetid, subnet_name, cidr, az, azid]
            for j in range(len(subnet_list)):
                row = chr(97 + j) + str(i + 20)
                infomation_dict[row] = subnet_list[j]

    def get_internat_gatewat_info(self):
        rows = ['a', 'b']
        igws_info = self.ec2_client.ec2_internet_gateways_describe(filters)
        for i in range(len(igws_info)):
            igw_info = igws_info[i]
            igw_id = igw_info['InternetGatewayId']
            igw_name = ''
            tags = igw_info['Tags']
            for tag in tags:
                if 'Name' == tag['Key']:
                    igw_name = tag['Value']
                    break
            row_igw_id = rows[0] + str(9 + i)
            row_igw_name = rows[1] + str(9 + i)
            infomation_dict[row_igw_id] = igw_id
            infomation_dict[row_igw_name] = igw_name

    def get_nat_gateway_info(self):
        ngws_info = self.ec2_client.ec2_nat_gateways_describe(filters)
        for i in range(len(ngws_info)):
            ngw_info = ngws_info[i]
            ngw_id = ngw_info['NatGatewayId']
            tags = ngw_info['Tags']
            ngw_name = ''
            for tag in tags:
                if 'Name' == tag['Key']:
                    ngw_name = tag['Value']
                    break
            subnet_id = ngw_info['SubnetId']
            ngw_public_ip = ngw_info['NatGatewayAddresses'][0]['PublicIp']
            ngw_private_ip = ngw_info['NatGatewayAddresses'][0]['PrivateIp']
            ngw_lists = [ngw_id, ngw_name, subnet_id, ngw_public_ip, ngw_private_ip]
            for j in range(len(ngw_lists)):
                row = chr(97 + j) + str(i + 14)
                infomation_dict[row] = ngw_lists[j]

    def get_route_tables_info(self):
        route_tables_info = self.ec2_client.ec2_route_tables_describe(filters)
        for i in range(len(route_tables_info)):
            route_table_id = route_tables_info[i]['RouteTableId']
            route_table_name = ''
            tags = route_tables_info[i]['Tags']
            for tag in tags:
                if tag['Key'] == "Name":
                    route_table_name = tag['Value']
                    break
            subnets_associations = ''
            associations = route_tables_info[i]['Associations']
            for association in associations:
                if 'SubnetId' in association.keys():
                    subnet_id = association['SubnetId']
                    subnets_associations = subnets_associations + subnet_id + ';'
            route_tables = route_tables_info[i]['Routes']
            for route_table in route_tables:
                destination = route_table['DestinationCidrBlock']
                route_table.pop('Origin')
                route_table.pop('State')
                route_table.pop('DestinationCidrBlock')
                target = ''
                for key in route_table.keys():
                    target = route_table[key]
                route_table_list = [route_table_id, route_table_name, subnets_associations, destination, target]
                for j in range(len(route_table_list)):
                    row = chr(97 + j) + str(i + 3)
                    route_tables_dict[row] = route_table_list[j]

    @staticmethod
    def store_excel():
        sheets = {'infomation': infomation_dict, 'instances': instances_dict, 'securitygroup': securitygroup_dict,
                  'EBS': ebs_dict, 'routetable': route_tables_dict}
        wb = load_workbook('template_resource.xlsx')
        for key in sheets.keys():
            ws = wb[key]
            content = sheets[key]
            for content_key in content.keys():
                ws[content_key] = content[content_key]
                bd = Side(style='thin', color='000000')
                row_num = ws.max_row
                column_num = ws.max_column
                for j in range(97, 97 + column_num):
                    for i in range(1, row_num + 1):
                        row = chr(j) + str(i)
                        p = ws[row]
                        p.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        wb.save('%saws_resource.xlsx' % filename)

    def main(self):
        self.get_vpc_info()
        self.get_internat_gatewat_info()
        self.get_nat_gateway_info()
        self.get_subnet_info()
        self.get_ec2_info()
        self.get_sg_info()
        self.get_volumes_info()
        self.get_route_tables_info()
        self.store_excel()


if __name__ == '__main__':
    app = GetInfo()
    app.main()
